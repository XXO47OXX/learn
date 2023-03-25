import asyncio
from bs4 import BeautifulSoup
from asyncio_runner import AsyncioRunner
from openpyxl import load_workbook, Workbook
import aiohttp
import re

cookie_baidu = 'BIDUPSID=46405A6E26362E579DBDAEFCE34CA742; PSTM=1676180166; BAIDUID=46405A6E26362E5723178FE076C37F29:FG=1; BAIDUID_BFESS=46405A6E26362E5723178FE076C37F29:FG=1; ZFY=u4gcIeCGgcXmvAsC7O:B3Zflel43KNv6lUxthvg1up7M:C; uc_login_unique=a90d4fc37f16f67809b2104e2c63cb8c; uc_recom_mark=cmVjb21tYXJrXzQ0MDMxMjQ2; BDUSS=hZfkFlbHNqRXBzVHJITExILTBmLXV6eUFncmk5MndyajNVMlIzV3RtTFUzU2RrSVFBQUFBJCQAAAAAAQAAAAEAAADjsT5kYWxlZGg4ODg4AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAANRQAGTUUABkUk; BDUSS_BFESS=hZfkFlbHNqRXBzVHJITExILTBmLXV6eUFncmk5MndyajNVMlIzV3RtTFUzU2RrSVFBQUFBJCQAAAAAAQAAAAEAAADjsT5kYWxlZGg4ODg4AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAANRQAGTUUABkUk; delPer=0; BD_CK_SAM=1; PSINO=7; H_PS_PSSID=38185_36560_38351_37862_38171_38290_38243_37923_38312_38284_26350_37957_22157_38281_37881; BD_UPN=12314753; BDORZ=B490B5EBF6F3CD402E515D22BCDA1598; BA_HECTOR=2g240h25802la5a50gaha4cf1i0tj041n; H_PS_645EC=42db5HuvZ/Yv9p3CQgq3E3hQF03WIhL9jwav6IyrgycR2bpzp/5ormzZAdc; baikeVisitId=860a1a71-9c4b-41fe-b865-264316049f9f; BD_HOME=1'


payload = {}
headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'Accept-Language': 'en-US,en;q=0.9',
    'Cache-Control': 'max-age=0',
    'Connection': 'keep-alive',
    'Cookie': cookie_baidu,
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'none',
    'Sec-Fetch-User': '?1',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36',
    'sec-ch-ua': '"Google Chrome";v="107", "Chromium";v="107", "Not=A?Brand";v="24"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"'
}

url = "https://www.baidu.com/s?wd="
ts = ['科普', '已更新']
async def fetch(url):
    async with aiohttp.ClientSession() as session:
        async with session.get(url) as response:
            return await response.text()
        
async def baidu_search(session, keyword, dest_filename, ws, wb, next ):
    total = 0
    if next == "":
        keyword_url = url + keyword
    else:
        keyword_url = url + next
    status = "屏蔽"
    try:
        async with session.get(keyword_url, data=payload, headers=headers) as response:
            if response.status == 200:
                lala = keyword_url
                # text = await response.text()
                text = await fetch(keyword_url)
                
                bs = BeautifulSoup(text, 'lxml')
                
                result = bs.select_one('.hint_PIwZX')
                
                if result:
                    rst = result.get_text()
                    print(rst)
                    fil = re.sub('\D',"",rst)
                    print(fil)
                    if int(fil) > 10000:
                        for title in bs.select('h3'):
                            tit = title.text
                            if any(t in tit for t in ts):
                                total +=1
                                status="OK"
                        ws.append([keyword,fil, status, total])
                        wb.save(filename=dest_filename)
                    else:
                        ws.append([keyword,fil, status, total])
                        wb.save(filename=dest_filename)
    except:
        print("Connection has been blocked by Baidu! Please wait...")
        await baidu_search(session, keyword, dest_filename, ws, wb, next )

async def main():
    workbook_domain_list = load_workbook(filename='Fil.xlsx', read_only=True)
    worksheet_domain_list = workbook_domain_list.active

    keyword_list =[]
    keywrd = ""
    dest_filename = 'result_baidu_search.xlsx'

    try:
        wb = load_workbook(filename=dest_filename)
    except FileNotFoundError as e:
        wb = Workbook()
    for row in worksheet_domain_list.rows:
        for cell in row:
            if cell.row == 1:
                continue
            if cell.column == 1:
                keywrd = str(cell.value)
        if keywrd:
            keyword_list.append(keywrd)

    ws = wb.active
    last_row = ws.max_row
    ws.title = "Baidu Search"


    if last_row == 1:
        ws.append(['Keywords','Results','Status','Total'])

    workbook_domain_list.close()
    next= ""
    if len(keyword_list) > 0:
        timeout_seconds = 5000
        session_timeout = aiohttp.ClientTimeout(total=None, sock_connect=timeout_seconds, sock_read=timeout_seconds)
        async with aiohttp.ClientSession(headers=headers, timeout=session_timeout) as session:
            runner = AsyncioRunner(10)
            for keyword in keyword_list:
                runner.add_task(baidu_search(session, keyword, dest_filename, ws, wb, next))

            while runner.running_task_count > 0:
                await asyncio.sleep(2)

if __name__ == '__main__':
    asyncio.run(main())
